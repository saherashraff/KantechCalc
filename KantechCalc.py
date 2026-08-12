#!/usr/bin/env python3
"""
Fire Suppression Quote Wizard – Python GUI Edition
Ported from the original Excel VBA macros.
Requires: openpyxl (pip install openpyxl)
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import math
import os
from typing import List, Dict, Optional, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not found. Excel export will be disabled.")
    print("Install it with: pip install openpyxl")

# ────────────────────────────────────────────────
# Engineering calculations (ported line‑by‑line)
# ────────────────────────────────────────────────

def ceil(x: float) -> int:
    return math.ceil(x)

def even(x: int) -> int:
    return x if x % 2 == 0 else x + 1

def lookup_table(fill_per_cyl: float, thresholds: List[tuple]):
    for limit, val in thresholds:
        if fill_per_cyl <= limit:
            return val
    return None

def calc_fm200(agent_qty: float, height: float) -> dict:
    T, U = 0.5621, max(height or 3, 0.5)
    K = ceil(agent_qty / 162)
    L = ceil(agent_qty / K)
    E = K * L
    M = lookup_table(L, [(8,8),(16,16),(32,32),(52,52),(95,106),(132,147),(162,180)])
    N = lookup_table(L, [(8,"303.205.015"),(16,"303.205.016"),(32,"303.205.017"),
                         (52,"303.205.018"),(95,"303.205.019"),(132,"303.205.020"),
                         (162,"303.205.021")])
    O = 314205322 if N in ("303.205.015","303.205.016","303.205.018") else 314205306
    P = 311700014 if N in ("303.205.015","303.205.016","303.205.017") else 311700006
    Q = ceil(K / 11)
    R = K - Q
    S = K - Q
    a_base = ceil(agent_qty / (T * U * 95))
    a = max(a_base * ceil(U / 4.87), 1)
    V = a if (agent_qty / a) <= 100 else a * 2
    W = max(even(ceil(agent_qty / (T * U * 25))), 2)
    X, Y = W // 2, W // 2
    return dict(K=K, L=L, E=E, M=M, N=N, O=O, P=P, Q=Q, R=R, S=S, V=V, W=W, X=X, Y=Y)

def calc_novec(agent_qty: float, height: float) -> dict:
    T, U = 0.5621, max(height or 3, 0.5)
    K = ceil(agent_qty / 187)
    L = ceil(agent_qty / K)
    E = K * L
    M = lookup_table(L, [(9.5,8),(19,16),(38,32),(62,52),(114,106),(158,147),(187,180)])
    N = lookup_table(L, [(9.5,"303.207.001"),(19,"303.207.002"),(38,"303.207.003"),
                         (62,"303.207.004"),(114,"303.207.005"),(158,"303.207.006"),
                         (187,"303.207.007")])
    O = 314207208 if N in ("303.207.001","303.207.002","303.207.004") else 314207321
    P = 311700014 if N in ("303.207.001","303.207.002","303.207.004") else 311700006
    Q = ceil(K / 11)
    R = K - Q
    S = K - Q
    a_base = ceil(agent_qty / (T * U * 96))
    a = max(a_base * ceil(U / 4.3), 1)
    V = a if (agent_qty / a) <= 136 else a * 2
    W = max(even(ceil(agent_qty / (T * U * 25))), 2)
    X, Y = W // 2, W // 2
    return dict(K=K, L=L, E=E, M=M, N=N, O=O, P=P, Q=Q, R=R, S=S, V=V, W=W, X=X, Y=Y)

# ────────────────────────────────────────────────
# Bill‑of‑materials builders
# ────────────────────────────────────────────────

def detection_full(D: str, C: int, X: int, Y: int) -> List[dict]:
    return [
        dict(D=D, F="4004-9302", H=1*C, K="", L="FA", M="Control Panel"),
        dict(D=D, F="batt",       H=2,     K="ch", L="FA", M="Local"),
        dict(D=D, F="Pro",        H=1,     K="ch", L="FA", M="Service"),
        dict(D=D, F="4098-5610",  H=X*C,  K="", L="FA", M="Field Device"),
        dict(D=D, F="4098-5261",  H=1,     K="ch", L="FA", M="Field Device"),
        dict(D=D, F="4098-5601",  H=Y*C,  K="", L="FA", M="Field Device"),
        dict(D=D, F="4098-5261",  H=1,     K="ch", L="FA", M="Field Device"),
        dict(D=D, F="2099-9149",  H=1*C, K="", L="FA", M="Field Device"),
        dict(D=D, F="2080-9067",  H=1*C, K="", L="FA", M="Field Device"),
        dict(D=D, F="4906-9127",  H=1*C, K="", L="FA", M="Notification"),
        dict(D=D, F="INT-GB6-24", H=1*C, K="", L="FA", M="Notification"),
    ]

def detection_co2(D: str, C: int=1) -> List[dict]:
    items = [
        dict(D=D, F="4004-9302", H=1, K="", L="FA", M="Control Panel"),
        dict(D=D, F="batt",       H=2, K="ch", L="FA", M="Local"),
        dict(D=D, F="pro",        H=1, K="ch", L="FA", M="Service"),
        dict(D=D, F="4098-5610",  H=1, K="", L="FA", M="Field Device"),
        dict(D=D, F="4098-5261",  H=1, K="ch", L="FA", M="Field Device"),
        dict(D=D, F="4098-5601",  H=1, K="", L="FA", M="Field Device"),
        dict(D=D, F="4098-5261",  H=1, K="ch", L="FA", M="Field Device"),
        dict(D=D, F="2099-9149",  H=1, K="", L="FA", M="Field Device"),
        dict(D=D, F="4906-9127",  H=1, K="", L="FA", M="Notification"),
        dict(D=D, F="INT-GB6-24", H=1, K="", L="FA", M="Notification"),
    ]
    if C > 1:
        for item in items:
            item["H"] *= C
    return items

def build_fm200_package(room: str, agent_qty: float, repeat: int, height: float) -> Optional[List[dict]]:
    c = calc_fm200(agent_qty, height)
    if c['M'] is None or c['N'] is None:
        return None
    D = "Tyco - Hygood (FM)"
    items = [
        dict(type="header2", text=room),
        dict(type="intro", text="FM-200 Systems Package includes below items :"),
    ]
    if c['K'] == 1:
        items += [
            dict(type="item", D=D, F=c['E'], H=1*repeat, K="", L="FF", M="Subtotal"),
            dict(type="item", D=D, F="300.205.001", H=c['E'], K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F=c['N'], H=1, K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F="304205030", H=1, K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F="304.205.006", H=1, K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F="302.205.025", H=1, K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F=c['O'], H=1, K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F=c['P'], H=1, K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F="310.205.224", H=c['V']*repeat, K="", L="FF", M="FM200"),
        ]
    else:
        items += [
            dict(type="item", D=D, F=c['E'], H=1*repeat, K="", L="FF", M="Subtotal"),
            dict(type="item", D=D, F="300.205.001", H=c['E'], K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F=c['N'], H=c['K'], K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F="304205030", H=c['Q'], K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F="304.205.006", H=c['K'], K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F="302.205.025", H=c['K'], K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F="304.209.004", H=c['R'], K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F="306.205.003", H=c['S'], K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F=c['O'], H=c['K'], K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F=c['P'], H=c['K'], K="ch", L="FF", M="FM200"),
            dict(type="item", D=D, F="310.205.224", H=c['V']*repeat, K="", L="FF", M="FM200"),
        ]
    items.append(dict(type="intro", text="Detection Package includes below items:"))
    for it in detection_full("Simplex", repeat, c['X'], c['Y']):
        items.append(dict(type="item", **it))
    return items

def build_novec_package(room: str, agent_qty: float, repeat: int, height: float) -> Optional[List[dict]]:
    c = calc_novec(agent_qty, height)
    if c['M'] is None or c['N'] is None:
        return None
    D = "Tyco - Hygood (SP)"
    items = [
        dict(type="header2", text=room),
        dict(type="intro", text="Novec Systems Package includes below items :"),
    ]
    if c['K'] == 1:
        items += [
            dict(type="item", D=D, F=c['E'], H=1*repeat, K="", L="FF", M="Subtotal"),
            dict(type="item", D=D, F="452011", H=c['E'], K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F=c['N'], H=1, K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F="304205030", H=1, K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F="304.205.006", H=1, K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F="302207021", H=1, K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F=c['O'], H=1, K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F=c['P'], H=1, K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F="310.207.220", H=c['V']*repeat, K="", L="FF", M="Novec"),
        ]
    else:
        items += [
            dict(type="item", D=D, F=c['E'], H=1*repeat, K="", L="FF", M="Subtotal"),
            dict(type="item", D=D, F="452011", H=c['E'], K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F=c['N'], H=c['K'], K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F="304205030", H=c['Q'], K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F="304.205.006", H=c['K'], K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F="302207021", H=c['K'], K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F="304.209.004", H=c['R'], K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F="306.205.003", H=c['S'], K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F=c['O'], H=c['K'], K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F=c['P'], H=c['K'], K="ch", L="FF", M="Novec"),
            dict(type="item", D=D, F="310.207.220", H=c['V']*repeat, K="", L="FF", M="Novec"),
        ]
    items.append(dict(type="intro", text="Detection Package includes below items:"))
    for it in detection_full("Simplex", repeat, c['X'], c['Y']):
        items.append(dict(type="item", **it))
    return items

def build_co2_ansul(room: str, cylinder_count: int=1) -> List[dict]:
    D = "Ansul"
    items = [
        dict(type="header2", text=room),
        dict(type="intro", text="CO2 Systems Package includes below items :"),
    ]
    for _ in range(cylinder_count):
        items += [
            dict(type="section", text="MASTER"),
            dict(type="item", D=D, F="451500", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="73327", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="428949", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="427082", H=1, K="ch", L="FF", M="CO2"),
            dict(type="section", text="SLAVE"),
            dict(type="item", D=D, F="451500", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="427082", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="426112", H=1, K="", L="FF", M="CO2"),
        ]
    items.append(dict(type="intro", text="Detection Package includes below items:"))
    for it in detection_co2("Simplex", cylinder_count):
        items.append(dict(type="item", **it))
    return items

def build_co2_hygood(room: str, cylinder_count: int=1) -> List[dict]:
    D = "Tyco - Hygood (LPG)"
    items = [
        dict(type="header2", text=room),
        dict(type="intro", text="CO2 Systems Package includes below items :"),
    ]
    for _ in range(cylinder_count):
        items += [
            dict(type="section", text="MASTER"),
            dict(type="item", D=D, F="71104006h", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="30521041", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="91104001", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="20006020", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="30027301", H=1, K="ch", L="FF", M="CO2"),
            dict(type="section", text="SLAVE"),
            dict(type="item", D=D, F="71104007h", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="30521041", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="30100102", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="30500041", H=1, K="ch", L="FF", M="CO2"),
            dict(type="item", D=D, F="30460002", H=1, K="", L="FF", M="CO2"),
        ]
    items.append(dict(type="intro", text="Detection Package includes below items:"))
    for it in detection_co2("Simplex", cylinder_count):
        items.append(dict(type="item", **it))
    return items

SYSTEM_HEADER1 = {
    "fm200": "Tyco / Hygood FM-200 Systems - UL Listed & FM Approved",
    "novec": "Hygood FK-5-1-12 Systems - UL Listed & FM Approved",
    "co2ansul": "Ansul CO2 - UL listed & FM Approved",
    "co2hygood": "Hygood (LPG) CO2 Systems - VdS Approved",
}

# ────────────────────────────────────────────────
# Main Application
# ────────────────────────────────────────────────

class QuoteWizardApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Fire Suppression Quote Wizard")
        self.root.geometry("1200x720")
        self.root.minsize(1000, 600)

        # Data
        self.quote: List[dict] = []            # original rows
        self.template_workbook = None
        self.template_filename = None

        # Toggle variables
        self.ws_toggle = tk.BooleanVar(value=True)
        self.ca_toggle = tk.BooleanVar(value=True)

        # Widget references
        self.tab_widgets = {}                  # dict keyed by system id
        self.calc_label = None

        self._create_layout()
        self._update_display()

    # ── Layout creation ────────────────────────
    def _create_layout(self):
        # Left panel
        left_frame = ttk.Frame(self.root, padding=10)
        self.left_frame = left_frame

        self.notebook = ttk.Notebook(left_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # FM‑200 tab
        fm200_tab = ttk.Frame(self.notebook)
        self._build_agent_tab(fm200_tab, "FM-200", "fm200")
        self.notebook.add(fm200_tab, text="FM-200")

        # Novec tab
        novec_tab = ttk.Frame(self.notebook)
        self._build_agent_tab(novec_tab, "Novec (FK-5-1-12)", "novec")
        self.notebook.add(novec_tab, text="Novec")

        # CO2 Ansul tab
        co2ansul_tab = ttk.Frame(self.notebook)
        self._build_co2_tab(co2ansul_tab, "Ansul CO2", "co2ansul")
        self.notebook.add(co2ansul_tab, text="CO2 – Ansul")

        # CO2 Hygood tab
        co2hygood_tab = ttk.Frame(self.notebook)
        self._build_co2_tab(co2hygood_tab, "Hygood (LPG) CO2", "co2hygood")
        self.notebook.add(co2hygood_tab, text="CO2 – Hygood")

        # Add button & preview
        self.add_btn = ttk.Button(left_frame, text="+ Add Package to Quote",
                                  command=self._on_add)
        self.add_btn.pack(fill=tk.X, pady=(10, 0))

        self.calc_label = tk.Label(left_frame, text="", anchor=tk.W, justify=tk.LEFT,
                                   bg="#f8f5ef", fg="#4a453d", font=("Courier", 10),
                                   padx=5, pady=5, relief=tk.GROOVE, wraplength=350)
        self.calc_label.pack(fill=tk.BOTH, pady=(10, 0))

        # Right panel
        right_frame = ttk.Frame(self.root, padding=10)
        self.right_frame = right_frame

        # Toolbar
        toolbar = ttk.Frame(right_frame)
        toolbar.pack(fill=tk.X, pady=(0, 10))

        self.export_btn = ttk.Button(toolbar, text="⬇ Export to Excel (Offer tab)",
                                     command=self._on_export)
        self.export_btn.pack(side=tk.LEFT, padx=2)

        self.load_btn = ttk.Button(toolbar, text="Use my template file…",
                                   command=self._on_load_template)
        self.load_btn.pack(side=tk.LEFT, padx=2)

        self.clear_btn = ttk.Button(toolbar, text="Clear Quote",
                                    command=self._on_clear)
        self.clear_btn.pack(side=tk.LEFT, padx=2)

        toggle_frame = ttk.Frame(toolbar)
        toggle_frame.pack(side=tk.RIGHT, padx=10)
        ttk.Checkbutton(toggle_frame, text="Warning Sign", variable=self.ws_toggle,
                        command=self._update_display).pack(side=tk.LEFT, padx=5)
        ttk.Checkbutton(toggle_frame, text="Class A Module", variable=self.ca_toggle,
                        command=self._update_display).pack(side=tk.LEFT, padx=5)

        self.template_status = ttk.Label(right_frame, text="", foreground="gray")
        self.template_status.pack(anchor=tk.W)

        # Treeview
        tree_frame = ttk.Frame(right_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("mfr", "part", "qty", "unit", "type", "category")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                                 selectmode="browse")
        for col in columns:
            self.tree.heading(col, text=col.capitalize())
        self.tree.column("mfr", width=120)
        self.tree.column("part", width=120)
        self.tree.column("qty", width=60)
        self.tree.column("unit", width=60)
        self.tree.column("type", width=60)
        self.tree.column("category", width=150)

        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Right‑click deletion
        self.tree.bind("<Button-3>", self._on_tree_right_click)
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="Delete Selected Row",
                                      command=self._delete_selected_row)

        # Note
        ttk.Label(right_frame, text="Right‑click a row to delete it. Toggled rows cannot be deleted individually.",
                  foreground="gray", font=("", 9)).pack(anchor=tk.W, pady=(5, 0))

        # Main paned window
        pw = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        pw.pack(fill=tk.BOTH, expand=True)
        pw.add(left_frame, weight=1)
        pw.add(right_frame, weight=2)

    def _build_agent_tab(self, parent: ttk.Frame, label: str, sys_id: str):
        frame = ttk.Frame(parent, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Room / System Name").pack(anchor=tk.W)
        room_entry = ttk.Entry(frame)
        room_entry.pack(fill=tk.X, pady=(0, 5))

        grid = ttk.Frame(frame)
        grid.pack(fill=tk.X, pady=5)

        ttk.Label(grid, text="Agent Quantity (kg)").grid(row=0, column=0, sticky=tk.W, padx=(0,5))
        qty_entry = ttk.Entry(grid, width=15)
        qty_entry.grid(row=1, column=0, sticky=tk.W, padx=(0,5))

        ttk.Label(grid, text="Repeated Times").grid(row=0, column=1, sticky=tk.W)
        repeat_entry = ttk.Entry(grid, width=10)
        repeat_entry.insert(0, "1")
        repeat_entry.grid(row=1, column=1, sticky=tk.W)

        ttk.Label(frame, text="Room Height (m)").pack(anchor=tk.W, pady=(5,0))
        height_entry = ttk.Entry(frame)
        height_entry.insert(0, "3")
        height_entry.pack(fill=tk.X)

        ttk.Label(frame, text=f"{label} sizing is automatic.",
                  foreground="gray", font=("", 9)).pack(anchor=tk.W, pady=(2,0))

        # Store widgets
        self.tab_widgets[sys_id] = {
            'room': room_entry, 'qty': qty_entry,
            'repeat': repeat_entry, 'height': height_entry
        }
        # Preview binding
        for w in [qty_entry, height_entry, repeat_entry]:
            w.bind("<KeyRelease>", lambda e, sid=sys_id: self._update_preview(sid))

    def _build_co2_tab(self, parent: ttk.Frame, label: str, sys_id: str):
        frame = ttk.Frame(parent, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Room / Package Description").pack(anchor=tk.W)
        room_entry = ttk.Entry(frame)
        room_entry.pack(fill=tk.X, pady=(0,5))

        ttk.Label(frame, text="Number of Cylinders (Master+Slave pairs)").pack(anchor=tk.W)
        cyl_entry = ttk.Entry(frame)
        cyl_entry.insert(0, "1")
        cyl_entry.pack(fill=tk.X)

        ttk.Label(frame, text=f"{label} package repeats Master/Slave BOM per cylinder.",
                  foreground="gray", font=("", 9)).pack(anchor=tk.W, pady=(2,0))

        self.tab_widgets[sys_id] = {'room': room_entry, 'cyl': cyl_entry}

    # ── Preview update ─────────────────────────
    def _update_preview(self, sys_id: str):
        if sys_id not in ("fm200", "novec"):
            self.calc_label.config(text="")
            return
        w = self.tab_widgets[sys_id]
        try:
            qty = float(w['qty'].get())
            if qty <= 0:
                raise ValueError
        except (ValueError, KeyError):
            self.calc_label.config(text="Enter an agent quantity to preview the sizing.")
            return

        try:
            height = float(w['height'].get()) or 3.0
            if height <= 0:
                height = 3.0
        except:
            height = 3.0

        calc_fn = calc_fm200 if sys_id == "fm200" else calc_novec
        c = calc_fn(qty, height)

        if c['M'] is None or c['N'] is None:
            self.calc_label.config(text="Cylinder fill exceeds maximum capacity.\nReduce agent quantity.")
            return

        txt = (f"Cylinder Qty: {c['K']}   Fill/Cyl: {c['L']} kg   Vol: {c['M']} L\n"
               f"P/N: {c['N']}   Label: {c['O']}   Bracket: {c['P']}\n"
               f"Elec Act: {c['Q']}   Pneum Act: {c['R']}   Hoses: {c['S']}\n"
               f"Nozzles: {c['V']}   Detectors: {c['W']} ({c['X']}H / {c['Y']}S)")
        self.calc_label.config(text=txt)

    # ── Add package ────────────────────────────
    def _on_add(self):
        tab_idx = self.notebook.select()
        tab_text = self.notebook.tab(tab_idx, "text")
        sys_map = {
            "FM-200": "fm200", "Novec": "novec",
            "CO2 – Ansul": "co2ansul", "CO2 – Hygood": "co2hygood"
        }
        sys_id = sys_map.get(tab_text)
        if not sys_id:
            return

        w = self.tab_widgets[sys_id]
        room = w['room'].get().strip()
        if not room:
            messagebox.showwarning("Input Error", "Please enter a room / package name.")
            return

        # Build package according to type
        try:
            if sys_id in ("fm200", "novec"):
                qty = float(w['qty'].get())
                if qty <= 0:
                    raise ValueError
                repeat = int(w['repeat'].get()) or 1
                height = float(w['height'].get()) or 3.0
                if height <= 0:
                    height = 3.0

                builder = build_fm200_package if sys_id == "fm200" else build_novec_package
                items = builder(room, qty, repeat, height)
                if items is None:
                    messagebox.showerror("Error", "Cylinder fill exceeds maximum capacity.\nReduce agent quantity.")
                    return

                # Reset fields
                w['room'].delete(0, tk.END)
                w['qty'].delete(0, tk.END)
                w['repeat'].delete(0, tk.END); w['repeat'].insert(0, "1")
                w['height'].delete(0, tk.END); w['height'].insert(0, "3")

            else:  # CO2
                cyl = int(w['cyl'].get()) or 1
                if cyl < 1:
                    raise ValueError
                builder = build_co2_ansul if sys_id == "co2ansul" else build_co2_hygood
                items = builder(room, cyl)
                w['room'].delete(0, tk.END)
                w['cyl'].delete(0, tk.END); w['cyl'].insert(0, "1")

        except ValueError:
            messagebox.showwarning("Input Error", "Please enter valid numerical values.")
            return

        # Ensure system header1 exists
        h1_key = sys_id
        if not any(r.get("type") == "header1" and r.get("system") == h1_key
                   for r in self.quote):
            self.quote.insert(0, dict(type="header1", system=h1_key,
                                      text=SYSTEM_HEADER1[h1_key]))

        # Tag and add
        for item in items:
            item["system"] = sys_id
        self.quote.extend(items)
        self._update_display()

    # ── Clear quote ────────────────────────────
    def _on_clear(self):
        if self.quote and not messagebox.askyesno("Clear Quote",
                                                  "Remove all items from the quote?"):
            return
        self.quote.clear()
        self._update_display()

    # ── Toggle application & display ───────────
    def _apply_toggles(self) -> List[dict]:
        """Return a list of rows with warning sign / class A module inserted."""
        rows = []
        ws = self.ws_toggle.get()
        ca = self.ca_toggle.get()
        for r in self.quote:
            rows.append(r)
            if r.get("type") == "item":
                if ws and r.get("F") == "INT-GB6-24":
                    rows.append(dict(type="item", system=r["system"],
                                     D="Simplex", F="4906-9101",
                                     H=1, K="", L="FF", M="Local",
                                     added=True))
                if ca and r.get("F") in ("Pro", "pro"):
                    rows.append(dict(type="item", system=r["system"],
                                     D="Simplex", F="4004-9864",
                                     H=1, K="", L="FA", M="Control Panel",
                                     added=True))
        return rows

    def _update_display(self, *args):
        """Refresh the treeview from the current quote + toggles."""
        self.tree.delete(*self.tree.get_children())

        rows = self._apply_toggles()
        # Tag definitions
        self.tree.tag_configure("header1", background="#1b2430",
                                foreground="white", font=("", 10, "bold"))
        self.tree.tag_configure("header2", background="#e7e0d2",
                                font=("", 10, "bold"))
        self.tree.tag_configure("intro", font=("", 10, "bold"))
        self.tree.tag_configure("section", foreground="gray",
                                font=("", 9, "italic"))
        self.tree.tag_configure("item", font=("", 10))

        for idx, row in enumerate(rows):
            if row.get("type") in ("header1", "header2", "intro", "section"):
                values = (row["text"], "", "", "", "", "")
            else:
                values = (row.get("D", ""), row.get("F", ""),
                          row.get("H", ""), row.get("K", ""),
                          row.get("L", ""), row.get("M", ""))
            # Store the original quote index if it's an original item, else None
            tags = [row.get("type", "item")]
            if not row.get("added"):
                # This row came from self.quote – find its index
                try:
                    qi = self.quote.index(row)
                    tags.append(f"orig_{qi}")
                except ValueError:
                    # shouldn't happen
                    pass
            self.tree.insert("", tk.END, iid=str(idx), values=values, tags=tuple(tags))

    # ── Row deletion ───────────────────────────
    def _on_tree_right_click(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)

    def _delete_selected_row(self):
        sel = self.tree.selection()
        if not sel:
            return
        item_iid = sel[0]
        item_tags = self.tree.item(item_iid, "tags")
        # Look for orig_N tag
        for tag in item_tags:
            if tag.startswith("orig_"):
                try:
                    orig_idx = int(tag.split("_")[1])
                    # Delete from self.quote
                    del self.quote[orig_idx]
                    self._update_display()
                    return
                except (IndexError, ValueError):
                    pass
        # If we get here, it's probably a toggled row
        messagebox.showinfo("Cannot delete",
                            "This row was added by a toggle (Warning Sign / Class A Module).\n"
                            "Uncheck the toggle to remove it.")

    # ── Template & Export ──────────────────────
    def _on_load_template(self):
        if not HAS_OPENPYXL:
            messagebox.showinfo("Missing Module",
                                "openpyxl is required for template loading.\n"
                                "Install with: pip install openpyxl")
            return
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm")])
        if not path:
            return
        try:
            self.template_workbook = openpyxl.load_workbook(path)
            self.template_filename = os.path.basename(path)
            self.template_status.config(
                text=f"Template loaded: {self.template_filename} – export will append into its 'Offer' tab."
            )
        except Exception as e:
            messagebox.showerror("Error", f"Could not read template file.\n{e}")

    def _on_export(self):
        if not self.quote:
            messagebox.showinfo("Empty Quote", "Add at least one package first.")
            return
        if not HAS_OPENPYXL:
            messagebox.showinfo("Missing Module",
                                "openpyxl is required for Excel export.\n"
                                "Install with: pip install openpyxl")
            return

        rows = self._apply_toggles()
        try:
            wb = self.template_workbook or openpyxl.Workbook()
            if "Offer" in wb.sheetnames:
                sheet = wb["Offer"]
                start_row = sheet.max_row + 2
            else:
                sheet = wb.create_sheet("Offer")
                start_row = 1

            r = start_row
            for row in rows:
                excel_row = sheet.row_dimensions[r]
                if row.get("type") == "header1":
                    cell = sheet.cell(row=r, column=7, value=row["text"])
                    cell.font = Font(bold=True, size=13, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1B2430", end_color="1B2430",
                                            fill_type="solid")
                    for col_idx in range(4, 14):
                        sheet.cell(row=r, column=col_idx).fill = PatternFill(
                            start_color="1B2430", end_color="1B2430", fill_type="solid")
                    excel_row.height = 22
                elif row.get("type") == "header2":
                    cell = sheet.cell(row=r, column=7, value=row["text"])
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="E7E0D2", end_color="E7E0D2",
                                            fill_type="solid")
                    excel_row.height = 18
                elif row.get("type") == "intro":
                    cell = sheet.cell(row=r, column=7, value=row["text"])
                    cell.font = Font(bold=True)
                elif row.get("type") == "section":
                    cell = sheet.cell(row=r, column=6, value=row["text"])
                    cell.font = Font(italic=True, color="999999")
                else:  # item
                    sheet.cell(row=r, column=4, value=row.get("D", ""))
                    sheet.cell(row=r, column=6, value=row.get("F", ""))
                    sheet.cell(row=r, column=8, value=row.get("H", ""))
                    sheet.cell(row=r, column=11, value=row.get("K", ""))
                    sheet.cell(row=r, column=12, value=row.get("L", ""))
                    sheet.cell(row=r, column=13, value=row.get("M", ""))
                r += 1

            # Column widths for a new workbook
            if not self.template_workbook:
                sheet.column_dimensions['D'].width = 22
                sheet.column_dimensions['F'].width = 16
                sheet.column_dimensions['G'].width = 40
                sheet.column_dimensions['H'].width = 8
                sheet.column_dimensions['K'].width = 8
                sheet.column_dimensions['L'].width = 6
                sheet.column_dimensions['M'].width = 16

            out_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=self.template_filename or "Fire_Suppression_Offer.xlsx"
            )
            if out_path:
                wb.save(out_path)
                messagebox.showinfo("Export Done", f"Offer sheet exported to {out_path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))


# ── Main entry point ──────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app = QuoteWizardApp(root)
    root.mainloop()
